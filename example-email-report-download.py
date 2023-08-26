# Automate email report Cell PLus

import string
import re
import smtplib
import time
import imaplib
import email
from email.header import decode_header
import traceback
import datetime
import dateutil.parser
from datetime import timedelta
import openpyxl

EMAIL_ACCOUNT = "example@example.com"
EMAIL_PWD = "password"
SMTP_SERVER = "imap.gmail.com"
SMTP_PORT = 993
OUTPUT_DIRECTORY_PATH = '/home/ubuntu/.example_reports/temp'

def copy_rows(src_sheet, dest_sheet, start_row, end_row):
    for row in src_sheet.iter_rows(min_row=start_row, max_row=end_row):
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        dest_sheet.append(new_row)

def read_email_from_gmail():
  try:
    mail = imaplib.IMAP4_SSL(SMTP_SERVER)

    typ, acct_details = mail.login(EMAIL_ACCOUNT,EMAIL_PWD)
    if typ != 'OK':
        print(f'Not able to sign in to {EMAIL_ACCOUNT} - {typ}')
        raise
    mail.select("Inbox")
    date = (datetime.date.today()-timedelta(days=0)).strftime("%d-%b-%Y")
    email_subject = 'Example Report'
    email_from = 'no-reply@example-report.com'
    search_string = f'(SENTON {date} FROM "{email_from}")'.format(date=date)
    (result, blocks) = mail.uid('search', None, search_string)

    num_messages = 0
    if result == "OK":
      for messages in blocks:
        for message in messages.split():
          ret, data = mail.uid('fetch', message, "(BODY.PEEK[])") # this keeps the message as unread
          if ret == "OK":
            raw = email.message_from_bytes(data[0][1])
            email_date = dateutil.parser.parse(decode_header(raw["Date"])[0][0]).strftime("%Y-%m-%d")
            if raw.get_content_maintype() != 'multipart':
              print('Error: no attachements found!')
              raise
            for part in raw.walk():
              if part.get_content_maintype() != 'multipart' and part.get('Content-Disposition') is not None:
                filename_bytes, charset = decode_header(part.get_filename())[0]
                filename = filename_bytes.decode(charset) if isinstance(filename_bytes, bytes) else filename_bytes      
                accepted_report_extensions = ['.csv', '.xls', '.xlsx']
                if any([x in filename for x in accepted_report_extensions]):
                  output_file_path = f'{OUTPUT_DIRECTORY_PATH}/{email_date}-{filename}' 
                  open(f'{output_file_path}', 'wb').write(part.get_payload(decode=True))
                  uid = message.split()[-1]
                  mail.uid('STORE', uid, '+X-GM-LABELS', 'uploaded/reports')
                  mail.uid('STORE',uid, '+FLAGS', '(\\Deleted)')
                  mail.expunge()
                  src_path = f'{OUTPUT_DIRECTORY_PATH}/{email_date}-{filename}'
                  dest_path = f'{OUTPUT_DIRECTORY_PATH}/{email_date}-{filename}'
                  src_book = openpyxl.load_workbook(src_path)
                  src_sheet = src_book['Auto Pilot Report']
                  dest_book = openpyxl.Workbook()
                  dest_sheet = dest_book.active
                  copy_rows(src_sheet, dest_sheet, 5, src_sheet.max_row)
                  dest_book.save(dest_path)
                  print(f'{dest_path}')
          else:
            print(f'Error fetching email {message}')
            raise
    else:
      print(f'Error searching {search_string} on {EMAIL_ACCOUNT}.')
      raise

  except Exception as e:
    traceback.print_exc() 
    print(str(e))

read_email_from_gmail()